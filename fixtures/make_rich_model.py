#!/usr/bin/env python3
"""Generate eval/fixtures/rich-model.xlsx — the PUBLISHABLE fidelity fixture.

The gs-model-edit task grades save fidelity against a real institutional
model, but that corpus file is proprietary and local-only. This fixture
carries the same *classes* of at-risk content — chart, comments, conditional
formatting, data validation, defined names, merges, freeze panes — in a
fully self-authored workbook that can live in the repo and back publishable
results. Regenerate after edits with:  python3 eval/fixtures/make_rich_model.py

Two classes stay covered ONLY by the private gs-model-edit corpus task:
pivot tables (openpyxl can't author them) and cell comments (openpyxl's
comment-part layout crashes ExcelJS's reconcile — and the app itself loads
through ExcelJS, so a fixture with them would break the eval run at open).
"""
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(__file__), "rich-model.xlsx")

wb = Workbook()

# --- Data sheet: quarterly revenue the chart references ---------------------
data = wb.active
data.title = "Data"
data.append(["Quarter", "Revenue", "COGS"])
quarters = ["Q1-24", "Q2-24", "Q3-24", "Q4-24", "Q1-25", "Q2-25", "Q3-25", "Q4-25"]
revenue = [1180, 1245, 1310, 1420, 1385, 1462, 1548, 1660]
for i, q in enumerate(quarters):
    data.append([q, revenue[i], round(revenue[i] * 0.62)])
for col in "BC":
    for row in range(2, 10):
        data[f"{col}{row}"].number_format = "#,##0"

# --- Model sheet: small P&L with formulas + every at-risk feature -----------
m = wb.create_sheet("Model")
m.merge_cells("A1:F1")
m["A1"] = "Acme Co — Operating Model (fixture)"
m["A1"].font = Font(bold=True, size=14, color="FFFFFF")
m["A1"].fill = PatternFill("solid", fgColor="1F4E79")
m["A1"].alignment = Alignment(horizontal="center")

m["A3"] = "Scenario"
m["B3"] = "Base"
dv = DataValidation(type="list", formula1='"Base,Bull,Bear"', allow_blank=False)
m.add_data_validation(dv)
dv.add(m["B3"])

m["A4"] = "Growth rate"
m["B4"] = 0.06
m["B4"].number_format = "0.0%"

headers = ["Line item"] + ["FY2024A", "FY2025A", "FY2026E", "FY2027E", "FY2028E"]
m.append([])  # row 5 spacer
m.append(headers)  # row 6
rows = [
    ("Revenue", 5155, 6055, "=C7*(1+GrowthRate*4)", "=D7*(1+GrowthRate*4)", "=E7*(1+GrowthRate*4)"),
    ("COGS", 3196, 3754, "=D7*0.62", "=E7*0.62", "=F7*0.62"),
    ("Gross profit", "=B7-B8", "=C7-C8", "=D7-D8", "=E7-E8", "=F7-F8"),
    ("Gross margin", "=B9/B7", "=C9/C7", "=D9/D7", "=E9/E7", "=F9/F7"),
]
for label, *vals in rows:
    m.append([label, *vals])
for col in range(2, 7):
    c = get_column_letter(col)
    for r in (7, 8, 9):
        m[f"{c}{r}"].number_format = "#,##0"
    m[f"{c}10"].number_format = "0.0%"
m["A6"].font = Font(bold=True)
for col in range(2, 7):
    m[f"{get_column_letter(col)}6"].font = Font(bold=True)
m.conditional_formatting.add(
    "B10:F10",
    ColorScaleRule(
        start_type="num", start_value=0.30, start_color="F8696B",
        end_type="num", end_value=0.45, end_color="63BE7B",
    ),
)
m.freeze_panes = "B7"
m.column_dimensions["A"].width = 24

wb.defined_names.add(DefinedName("GrowthRate", attr_text="Model!$B$4"))

chart = BarChart()
chart.title = "Quarterly revenue"
chart.add_data(Reference(data, min_col=2, min_row=1, max_row=9), titles_from_data=True)
chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=9))
m.add_chart(chart, "H3")

wb.save(OUT)

# --- inject parts OUTSIDE openpyxl's model ----------------------------------
# openpyxl (and any library-based rewrite) silently DROPS zip parts it
# doesn't model on load→save; GridPath's surgical patcher copies untouched
# parts byte-identical. Without these the fixture only contains features
# both pipelines round-trip, and the fidelity axis can't discriminate —
# both harnesses scored 7/7 on v1 of this fixture for exactly that reason.
import zipfile

CUSTOM_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<gridpathEval xmlns="urn:gridpath:eval"><provenance>fixture v2</provenance></gridpathEval>'
)
PRINTER_BIN = bytes(range(64))  # opaque payload; content is irrelevant, survival is the test

tmp = OUT + ".tmp"
with zipfile.ZipFile(OUT) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename == "[Content_Types].xml":
            data = data.replace(
                b"</Types>",
                b'<Default Extension="bin" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings"/>'
                b'<Override PartName="/customXml/item1.xml" ContentType="application/xml"/></Types>',
            )
        zout.writestr(item, data)
    zout.writestr("customXml/item1.xml", CUSTOM_XML)
    zout.writestr("xl/printerSettings/printerSettings1.bin", PRINTER_BIN)
os.replace(tmp, OUT)
print(f"wrote {OUT} (with injected customXml + printerSettings)")
